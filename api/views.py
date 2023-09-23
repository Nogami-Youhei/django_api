from django.shortcuts import render, redirect
from .forms import ReportForm, SearchForm, SignupForm, LoginForm, CustomUserChangeForm, CustomSetPasswordForm
from .models import Report, Category
from django.http import HttpResponse, FileResponse, JsonResponse
from django.utils import timezone
import datetime
import pytz
from django.db.models import Q
from django.core.paginator import Paginator
from django.contrib.auth import login, logout
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
import os
from pathlib import Path
import shutil
import pandas as pd
import openpyxl
from openpyxl.styles import Font
import matplotlib.pyplot as plt
import japanize_matplotlib
import matplotlib.dates as mdates


THIS_FOLDER = Path(__file__).parent.resolve()

def signup_view(request):
    if request.method == 'POST':
        form = SignupForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            return redirect(to='/api/')
        
    else:
        form = SignupForm()
    param = {
        'title': 'ユーザー登録',
        'form': form,
        'url': '/api/signup/'
        }
    return render(request, 'api/signup.html', param)


def login_view(request):
    if request.method == 'POST':
        next = request.POST.get('next')
        form = LoginForm(request, data=request.POST)

        if form.is_valid():
            user = form.get_user()

            if user:
                login(request, user)
                if next == 'None':
                    return redirect(to='/api/')
                else:
                    return redirect(to=next)

    else:
        form = LoginForm()
        next = request.GET.get('next')

    params = {
        'form': form,
        'next': next,
    }

    return render(request, 'api/login.html', params)


def logout_view(request):
    logout(request)
    return redirect('login')


@login_required
def rename_view(request):
    if request.method == 'POST':
        form = CustomUserChangeForm(request.POST, instance=request.user)
        if form.is_valid():
            form.save()
            return redirect(to='/api/')
        
    else:
        form = CustomUserChangeForm(instance=request.user)
    
    form.fields.pop('password')

    param = {
        'title': 'ユーザー名変更',
        'form': form,
        'url': '/api/rename/'
    }
    return render(request, 'api/signup.html', param)


@login_required
def setpassword_view(request):
    if request.method == 'POST':
        form = CustomSetPasswordForm(user=request.user, data=request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)

            return redirect(to='/api/')
        
    else:
        form = CustomSetPasswordForm(user=request.user)
    
    param = {
        'title': 'パスワード変更',
        'form': form,
        'url': '/api/setpassword/',
    }
    return render(request, 'api/signup.html', param)


@login_required
def index(request):

    user = request.user.last_name + ' ' + request.user.first_name

    if request.method == 'POST':
        form = ReportForm(request.POST)

        if form.is_valid():
            report = form.save()
            smile_id = request.user.username
            last_name = request.user.last_name
            first_name = request.user.first_name
            report.name = last_name + ' ' + first_name
            report.smile_id = smile_id
            report.author = request.user
            report.save()
            
            return HttpResponse('success')
        
        else:
            search_form = SearchForm()
            params = {
                'user': user,
                'report_form': form,
                'search_form': search_form,
            }
            return render(request, 'api/index.html', params)
    
    search_form = SearchForm()
    report_form = ReportForm()
    
    params = {
        'user': user,
        'report_form': report_form,
        'search_form': search_form,
    }

    return render(request, 'api/index.html', params)


def search_result(keywords, categories, start_date, end_date):
    items = Report.objects.all()
    if keywords:
        keywords = keywords.split()
        query_filter = (Q(name__icontains=keywords[0]) | Q(title__icontains=keywords[0]) | Q(abstract__icontains=keywords[0]))

        for keyword in keywords:
            query_filter &= (Q(name__icontains=keyword) | Q(title__icontains=keyword) | Q(abstract__icontains=keyword))
        
        items = items.filter(query_filter)
    if categories:
        items = items.filter(categories__in=categories)
    if start_date and end_date:
        timezone = pytz.timezone('Asia/Tokyo')

        # datetime.date オブジェクトを datetime.datetime オブジェクトに変換してタイムゾーン情報を追加
        start_date = datetime.datetime.combine(start_date, datetime.time())
        end_date = datetime.datetime.combine(end_date, datetime.time(23, 59, 59, 999999))

        start_date = timezone.localize(start_date)
        end_date = timezone.localize(end_date)
        items = items.filter(datetime__range=(start_date, end_date))

    items = items.distinct()
    return items


@login_required
def search(request):

    form = SearchForm(request.POST)

    if form.is_valid():
        keywords = form.cleaned_data.get('keyword')
        categories = form.cleaned_data.get('categories')
        start_date = form.cleaned_data.get('start_date')
        end_date = form.cleaned_data.get('end_date')
        sort = form.cleaned_data.get('sort')

        items = search_result(keywords, categories, start_date, end_date)

        if sort == '1':
            items = items.order_by('datetime')

        elif sort == '2':
            items = items.order_by('title')

        elif sort == '3':
            items = items.order_by('-readers_number')
            
        paginator = Paginator(items, 8)
        number = request.POST.get('p', 1)
        items = paginator.page(number)

        author = request.user
        params = {
            'items': items,
            'author': author,
        }

        return render(request, 'api/items.html', params)
    
    else:
        return HttpResponse('error')


@login_required
def detail(request):
    id = request.POST.get('detail')
    report = Report.objects.get(id=id)
    categories = report.categories.all()
    users = report.readers.all()
    categories = [category.name for category in categories]
    users = [user.last_name + ' ' + user.first_name for user in users]
    readers_number = report.readers_number
    params = {
        'report': report,
        'categories': categories,
        'users': users,
        'readers_number': readers_number,
    }

    return render(request, 'api/detail.html', params)


@login_required
def delete(request):
    checked_list = request.POST.get('checked_list')
    print(type(checked_list))
    checked_list = [int(i) for i in checked_list.split(',')]
    print(checked_list)
    report = Report.objects.filter(id__in=checked_list)
    report.delete()
    return HttpResponse('delete')


@login_required
def update(request):
    if request.method == 'POST':
        id = request.POST.get('id')
        report = Report.objects.get(id=id)
        form = ReportForm(request.POST, instance=report)
        if form.is_valid():
            form.save()
            return HttpResponse('success')
    
    id = request.GET.get('id')
    report = Report.objects.get(id=id)
    report_form = ReportForm(instance=report)
    params = {
        'report_form': report_form,
        'id': id,
    }
    return render(request, 'api/update.html', params)


@login_required
def box(request):
    id = request.POST.get('box')
    report = Report.objects.get(id=id)
    user = request.user
    report.readers.add(user)
    readers = report.readers.all()
    report.readers_number = readers.count()
    report.save()
    result = {
        'status': 'success',
        'url': report.box_url,
    }
    
    return JsonResponse(result)


@login_required
def output(request):
    
    form = SearchForm(request.POST)

    if form.is_valid():
        keywords = form.cleaned_data.get('keyword')
        categories = form.cleaned_data.get('categories')
        start_date = form.cleaned_data.get('start_date')
        end_date = form.cleaned_data.get('end_date')
        sort = form.cleaned_data.get('sort')

        items = search_result(keywords, categories, start_date, end_date)

        if sort == '1':
            items = items.order_by('datetime')

        elif sort == '2':
            items = items.order_by('title')

        elif sort == '3':
            items = items.order_by('-readers_number')

        temp_dir = THIS_FOLDER.joinpath('temp')
        shutil.rmtree(temp_dir)
        os.makedirs(temp_dir, exist_ok=True)
        file_dir = temp_dir.joinpath('result.xlsx')
        df = pd.DataFrame(list(items.values()))
        df['datetime'] = df['datetime'].dt.tz_localize(None)+ pd.Timedelta(hours=9)
        df.columns = ['ID', '氏名', 'Smile_ID', 'タイトル', '要約', 'BOX_URL', '登録日時', '閲覧者数', '登録者ID']
        df.to_excel(file_dir)

        wb = openpyxl.load_workbook(file_dir)
        ws = wb.active
        ws.auto_filter.ref = "A1:J1"
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 50
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 20
        wrap_text = openpyxl.styles.Alignment(wrapText=True, vertical='center')

        for row in range(2, ws.max_row+1):
            ws.row_dimensions[row].height = 100

        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = wrap_text

        for cell in ws[f'F2:F{ws.max_row}']:
            cell[0].alignment = openpyxl.styles.Alignment(wrapText=True, vertical='top')

        for cell in ws['A']:
            cell.alignment = openpyxl.styles.Alignment(wrapText=True, horizontal='center', vertical='center')

        for cell in ws['B']:
            cell.alignment = openpyxl.styles.Alignment(wrapText=True, horizontal='center', vertical='center')

        for cell in ws[f'G2:G{ws.max_row}']:
            cell[0].hyperlink = cell[0].value
            cell[0].font = Font(color="0000FF", underline="single")

        wb.save(file_dir)
        filename, filepath = f'{keywords}.xlsx', file_dir
        return FileResponse(open(filepath, 'rb'), as_attachment=True, filename=filename)


@login_required
def analysis(request):
    
    items = Report.objects.all()

    df = pd.DataFrame(list(items.values()))
    df.index = df['datetime'].dt.tz_localize(None)+ pd.Timedelta(hours=9)

    df = df.resample('T').count()

    plt.plot(df.index, df['id'])
    plt.xlabel('登録日時')
    plt.ylabel('登録件数')
    plt.gca().xaxis.set_major_formatter(mdates.DateFormatter('%H:%M'))
    path_plot = THIS_FOLDER / 'static' / 'api' / 'img' / 'plot.png'
    plt.savefig(path_plot)
    plt.close()

    sizes = []
    labels = []

    for i in range(1, Category.objects.count()+1):
        sizes.append(Category(id=i).reports.count())

    for category in (Category.objects.all().order_by('name')):
        labels.append(category.name)

    colors = ['#ff9999', '#66b3ff', '#99ff99', '#ffcc99', '#c2c2f0', '#ffb3e6', '#ffb366']

    fig1, ax1 = plt.subplots()
    wedges, texts, autotexts = ax1.pie(sizes, labels=labels, colors=colors, autopct='%1.1f%%', shadow=True, startangle=140)

    path_plot = THIS_FOLDER / 'static' / 'api' / 'img' / 'pie.png'
    plt.savefig(path_plot)
    plt.close()  

    params = {
        'timestamp': timezone.now().timestamp(),
    }
    
    return render(request, 'api/analysis.html', params)



from api.models import Report
from api.serializers import ReportSerializer, AuthorSerializer, CategorySerializer
from rest_framework import generics
from rest_framework import permissions
from api.permissions import IsAuthorOrReadOnly
from rest_framework import viewsets

User = get_user_model()

class ReportViewSet(viewsets.ModelViewSet):
    queryset = Report.objects.all()
    serializer_class = ReportSerializer
    permission_classes = [permissions.IsAuthenticated, IsAuthorOrReadOnly]

    def perform_create(self, serializer):
        user = self.request.user

        serializer.save(
            name=user.last_name + ' ' + user.first_name,
            smile_id=user.username,
            author=user,
        )


class AuthorViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = User.objects.all()
    serializer_class = AuthorSerializer

    permission_classes = [permissions.IsAuthenticated]


class CategoryViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = Category.objects.all()
    serializer_class = CategorySerializer

    permission_classes = [permissions.IsAuthenticated]


from rest_framework.authtoken.models import Token

@login_required
def token_view(request):
    token, created = Token.objects.get_or_create(user=request.user)
    return HttpResponse(token)


from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework.reverse import reverse


@api_view(['GET'])
def api_root(request, format=None):
    return Response({
        'authors': reverse('author-list', request=request, format=format),
        'reports': reverse('report-list', request=request, format=format),
        'categories': reverse('category-list', request=request, format=format),
    })